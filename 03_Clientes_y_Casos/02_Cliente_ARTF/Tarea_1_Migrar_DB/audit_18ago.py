import openpyxl
import requests

ENV_PATH = ".env"
def load_env(path):
    env = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env

ENV = load_env(ENV_PATH)
SUPABASE_URL = ENV["SUPABASE_URL"]
SERVICE_KEY = ENV["SUPABASE_SERVICE_ROLE_KEY"]
HEADERS = {"apikey": SERVICE_KEY, "Authorization": f"Bearer {SERVICE_KEY}"}

def fetch_all(table, select):
    out, offset, page = [], 0, 1000
    while True:
        r = requests.get(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS,
                          params={"select": select, "limit": page, "offset": offset}, timeout=60)
        r.raise_for_status()
        batch = r.json()
        out.extend(batch)
        if len(batch) < page:
            break
        offset += page
    return out

XLSX_PATH = "CRM - Leads Campaña 1 Reconexión Financiera (18 de agosto 6 pm).xlsx"
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

print("==== 1. Activity Log ====")
ws_log = wb["Activity Log"]
rows = ws_log.iter_rows(values_only=True)
header_log = list(next(rows))
sheet_logs = [row for row in rows if any(row)]
print(f"Filas en Sheet Activity Log: {len(sheet_logs)}")
sup_logs = fetch_all("activity_log", "id,created_at,cliente_id,gestion_lead_id")
print(f"Filas en Supabase activity_log: {len(sup_logs)}")

print("\n==== 2. 133 filas sin ManyChat ID ====")
ws_crm = wb["CRM"]
crm_rows = ws_crm.iter_rows(values_only=True)
header_crm = list(next(crm_rows))
leads_no_mcid = []
crm_dicts = []
for row in crm_rows:
    d = {header_crm[i]: row[i] for i in range(32)}
    if all(v in (None, "") for v in d.values()):
        continue
    crm_dicts.append(d)
    if not d.get("ManyChat ID"):
        leads_no_mcid.append(d)

sup_clientes = fetch_all("clientes", "id,manychat_id,ig_handle,nombre")

matched_by_ig = []
matched_by_name = []
for ld in leads_no_mcid:
    ig = ld.get("IG Handle")
    name = ld.get("Nombre")
    found = False
    if ig:
        ig_norm = str(ig).replace("@", "").lower().strip()
        for c in sup_clientes:
            if c.get("ig_handle") == ig_norm:
                matched_by_ig.append(ld)
                found = True
                break
        if found:
            continue
    if name:
        name_norm = str(name).lower().strip()
        for c in sup_clientes:
            if c.get("nombre") and str(c.get("nombre")).lower().strip() == name_norm:
                matched_by_name.append(ld)
                found = True
                break

print(f"Match por IG Handle: {len(matched_by_ig)}")
print(f"Match por Nombre: {len(matched_by_name)}")
irresolubles = len(leads_no_mcid) - len(matched_by_ig) - len(matched_by_name)
print(f"Irresolubles con certeza razonable: {irresolubles}")

print("\n==== 3. El lado financiero ====")
sheet_ganados = [d for d in crm_dicts if str(d.get("Estado")).strip() == "Ganado"]
sup_ventas = fetch_all("ventas", "id,gestion_lead_id")
print(f"Leads 'Ganado' en Sheet: {len(sheet_ganados)}")
print(f"Filas en Supabase 'ventas': {len(sup_ventas)}")
print(f"Diferencia Ganados Sheet vs Ventas DB: {len(sheet_ganados)} vs {len(sup_ventas)}")

sup_pagos = fetch_all("pagos_cuotas", "id")
print(f"Filas en Supabase 'pagos_cuotas': {len(sup_pagos)}")

# Check missing ventas
sup_leads = fetch_all("gestion_leads", "id,cliente_id")
sup_leads_dict = {lead["id"]: lead["cliente_id"] for lead in sup_leads}
cli_dict = {c["id"]: c["manychat_id"] for c in sup_clientes}
mcids_with_venta = [cli_dict.get(sup_leads_dict.get(v["gestion_lead_id"])) for v in sup_ventas]
missing_ventas = [d for d in sheet_ganados if str(d.get("ManyChat ID")) not in mcids_with_venta]
if missing_ventas:
    print(f"Faltan en DB {len(missing_ventas)} ganados:")
    for m in missing_ventas[:10]:
        print(f" - Nombre: {m.get('Nombre')}, ManyChat: {m.get('ManyChat ID')}")
else:
    print("No faltan ganados en Supabase.")

print("\n==== 5. 15 conflictos sin resolver ====")
sup_reuniones = fetch_all("reuniones", "id,gestion_lead_id,estado,cliente_id,fecha_programada")
reuniones_realizadas = [r for r in sup_reuniones if r["estado"] == "realizada"]
mcid_to_dict = {str(d.get("ManyChat ID")): d for d in crm_dicts if d.get("ManyChat ID")}

conflictos = []
for r in reuniones_realizadas:
    cli_id = sup_leads_dict.get(r["gestion_lead_id"])
    mcid = cli_dict.get(cli_id)
    if mcid and mcid_to_dict.get(mcid) and str(mcid_to_dict[mcid].get("Estado")).strip() == "No show":
        d = mcid_to_dict[mcid]
        conflictos.append({
            "nombre": d.get("Nombre"),
            "mcid": mcid,
            "notas": d.get("Notas"),
            "fecha_reunion_sheet": d.get("Fecha Llamada Programada"),
            "fecha_reunion_db": r.get("fecha_programada")
        })

print(f"Encontrados {len(conflictos)} conflictos 'realizada' vs 'No show'")
for c in conflictos:
    print(f" - Nombre: {c['nombre']}, MCID: {c['mcid']}, Notas: {c['notas']}")
